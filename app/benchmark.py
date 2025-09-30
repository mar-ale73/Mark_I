# app/benchmark.py
from __future__ import annotations

import argparse
from pathlib import Path
from typing import Optional

import pandas as pd
import yaml


def _read_yaml(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def _ensure_parent(p: Path) -> None:
    p.parent.mkdir(parents=True, exist_ok=True)


def _write_separate_excel(df: pd.DataFrame, out_path: Path, sheet_name: str = "Comparación_Modelos") -> None:
    _ensure_parent(out_path)
    with pd.ExcelWriter(out_path, engine="openpyxl", mode="w") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)


def _append_sheet_replace(out_report: Path, df: pd.DataFrame, sheet_name: str = "Comparación_Modelos") -> None:
    """
    Agrega (o reemplaza si existe) la hoja `sheet_name` al Excel de reportes principal.
    No borra otras hojas.
    """
    _ensure_parent(out_report)

    if out_report.exists():
        try:
            from openpyxl import load_workbook
            wb = load_workbook(out_report)
            if sheet_name in wb.sheetnames:
                std = wb[sheet_name]
                wb.remove(std)
                wb.save(out_report)
        except Exception as e:
            print(f"⚠️ No se pudo preparar el Excel para reemplazo de hoja: {e}")

        with pd.ExcelWriter(out_report, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        with pd.ExcelWriter(out_report, engine="openpyxl", mode="w") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)


def main(argv: Optional[list[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Runner de Benchmark (standalone).")
    parser.add_argument("--config", default="utils/config.yaml", help="Ruta al YAML de configuración.")
    parser.add_argument("--out", default=None, help="Excel aparte para el benchmark.")
    parser.add_argument(
        "--no-append-report",
        action="store_true",
        help="No agrega la hoja 'Comparación_Modelos' al Excel principal.",
    )
    args = parser.parse_args(argv)

    cfg_path = Path(args.config)
    cfg = _read_yaml(cfg_path)

    try:
        from modelos.evaluacion_modelos import ejecutar_benchmark
    except Exception as e:
        print("❌ No se pudo importar 'ejecutar_benchmark' desde modelos.evaluacion_modelos.")
        print(f"   Detalle: {e}")
        return 1

    try:
        df_resultados = ejecutar_benchmark(cfg)
        if not isinstance(df_resultados, pd.DataFrame) or df_resultados.empty:
            print("⚠️ Benchmark no devolvió resultados.")
            return 2
    except Exception as e:
        print("❌ Error ejecutando el benchmark.")
        print(f"   Detalle: {e}")
        return 1

    out_bench = Path(args.out) if args.out else Path(cfg.get("ruta_reporte_benchmark", "outputs/benchmark_resultados.xlsx"))
    try:
        _write_separate_excel(df_resultados, out_bench, sheet_name="Comparación_Modelos")
        print(f"✅ Benchmark finalizado. Resultado escrito en: {out_bench}")
    except Exception as e:
        print("❌ No se pudo escribir el Excel de resultados del benchmark.")
        print(f"   Detalle: {e}")
        return 1

    if not args.no_append_report:
        try:
            out_report = Path(cfg.get("ruta_reporte", "outputs/reporte_inversion.xlsx"))
            _append_sheet_replace(out_report, df_resultados, sheet_name="Comparación_Modelos")
            print(f"📄 Hoja 'Comparación_Modelos' agregada a: {out_report}")
        except Exception as e:
            print("⚠️ No se pudo agregar la hoja 'Comparación_Modelos' al Excel principal.")
            print(f"   Detalle: {e}")

    try:
        cols_show = [
            c for c in [
                "Modelo","Horizonte","MAE","RMSE","MAPE_%","R2",
                "Accuracy_direccional","Sortino","Profit_Factor","MaxDD_%","Ret_Acumulado_%","Trades"
            ] if c in df_resultados.columns
        ]
        print("\n— Resumen (primeras filas) —")
        print(df_resultados[cols_show].head(10).to_string(index=False))
    except Exception:
        pass

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
